# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Alignment
#from .items import AnimateItem
from scrapy.pipelines.images import ImagesPipeline
import scrapy
from datetime import datetime
from PIL import Image
from urllib.parse import urlencode

class ResultManager(object):
    result_row_height = 50
    thumbnail_column_width = 50

    columns = {
        'datetime': 1,
        'name': 2,
        'sku': 3,
        'url': 4,
        'price': 5,
        'point': 6,
        'image1_capture': 7,
        'description': 8,
        'sale_start': 9,
        'sale_status': 10,
        'bonus': 11,
        'purchase_limit': 12,
        'site': 13,
        'genre':14,
        'image_url_1': 15,
        'image_url_2': 16,
        'image_url_3': 17,
        'image_url_4': 18,
        'image_url_5': 19,
        'image_url_6': 20,
        'image_url_7': 21,
        'image_url_8': 22,
        'image_url_9': 23,
        'image_url_10': 24
    }
    start_row = 2

    def __init__(self, sheet):
        self.sheet = sheet

    def get_target_row(self, sku):
        i = self.start_row
        while True:
            current_sku = self.sheet.cell(row = i, column = self.columns['sku']).value
            if not current_sku or current_sku == sku:
                return i
            i += 1
    
    def check_if_already_exists(self, item):
        for i in range(self.start_row, self.sheet.max_row + 1):
            cell = self.sheet.cell(row = i, column = self.columns['sku'])
            if cell.value == item['sku']:
                return True
        return False

    def write_cell(self, r, c, value):
        if type(value) == list:
            value = '\n›'.join(value)
        cell = self.sheet.cell(row = r, column = c)
        cell.alignment = Alignment(wrapText=True)
        cell.value = value
    
    def embed_thumbnail(self, item):
        row = self.get_target_row(item['sku'])
        img = openpyxl.drawing.image.Image(item["thumbnail_path"])
        cell = self.sheet.cell(row = row, column = self.columns["image1_capture"])
        img.anchor = cell.coordinate
        self.sheet.add_image(img)

    def record_item(self, item):
        # change row heigh
        target_row = self.get_target_row(item['sku'])
        self.sheet.row_dimensions[target_row].height = self.result_row_height
        for col_name, col_num in self.columns.items():
            if 'image_url_' in col_name:
                image_no = int(col_name.replace('image_url_', ''))
                if len(item['image_urls']) >= image_no:
                    self.write_cell(target_row, col_num, item['image_urls'][image_no - 1])
            else:
                try:
                    value = item[col_name]
                except KeyError: # if key not exists
                    pass # do nothing
                else:
                    self.write_cell(target_row, col_num, value)

class SearchQuery(object):
    def __str__(self):
        return self.urlparams
    categories = {
        'keyname': 'spc',
        'related_keyname': 'scc',
        'items': {
            'グッズ': {
                'value': "1",
                'related': {
                    'ストラップ・キーホルダー': '10',
                    'バッチ類': '11',
                    'バッグ・財布類': '12',
                    'スマホ雑貨': '13',
                    '収納商品': '14',
                    'アパレル・キャラクターアイテム': '15',
                    'タオル・ハンカチ類': '16',
                    '抱き枕・他インテリア': '17',
                    '文具・デスク用品': '18',
                    'コスメ関連': '19',
                    'アクセサリー': '20',
                    'キッチン雑貨': '21',
                    'TCGグッズ': '22',
                    'ポスター・タペストリー類': '23',
                    'ブロマイド・フォトアルバム': '24',
                    '食品・お菓子': '25',
                    'その他グッズ': '26'
                }
            },
            '映像': {
                'value': '2',
                'related': {
                    'DVD': '27',
                    'Blu-ray': '28'
                }
            },
            '音楽': {
                'value': '3',
                'related': {
                    'アルバム': '30',
                    '主題歌': '31',
                    'ドラマCD': '32',
                    'サウンドトラック': '33',
                    'キャラクターソング': '34',
                    'マキシシングル': '35',
                    'ラジオCD・DJCD': '36',
                    'その他音楽': '37',
                    'イヤホン・ヘッドホン': '75',
                    'データ販売': '82'
                }
            },
            '書籍': {
                'value': '4',
                'related': {
                    'コミック': '38',
                    '小説': '39',
                    'その他書籍': '40',
                    '雑誌': '41',
                    'ビジュアルファンブック': '42',
                    '写真集': '43',
                    'イラスト集・画集・原画集': '44',
                    '設定集・資料集': '45'        
                }
            },
            'フィギュア': {
                'value': '5',
                'related': {
                    '男性キャラクター': '46',
                    '女性キャラクター': '47',
                    'その他': '48',
                    '箱買いフィギュア': '49'
                }
            },
            'ゲーム': {
                'value': '6',
                'related': {
                    'Vita': '50',
                    'Win': '51',
                    'PS4': '52',
                    '3DS': '53',
                    'PSP': '54',
                    'PS3': '55',
                    'WiiU': '56',
                    'Switch': '57',
                    'その他ゲーム': '58',
                    '攻略本': '59'
                }
            },
            'チケット': {
                'value': '7',
                'related': {
                    'ライブ・コンサート': '60',
                    'イベント・原画展': '61',
                    '舞台': '62',
                    '映画': '63'   
                }
            },
            '同人': {
                'value': '8',
                'related': {
                    '同人誌': '64',
                    '同人CD': '65',
                    '同人グッズ': '66',
                    '同人ソフト': '67',
                    '同人DVD': '68'   
                }
            },
            'その他':{
                'value': '9',
                'related': {
                    'カレンダー': '69',
                    '図書カード・アニメイトコイン': '70',
                    'パンフレット': '71',
                    '画材': '72',
                    'コスプレ': '73',
                    'アウトレット': '74'
                }
            },
            '書泉': {
                'value': '76',
                'related': {
                    '鉄道・バス': '77',
                    'アイドル・グラビア': '78',
                    'プロレス・格闘技': '79',
                    'ミリタリー': '80',
                    'その他': '81'
                }
            }
        }
    }
    sonota = {
        'keyname': 'nd',
        'new_arrival': '1',
        'bonus': '5',
        'reservation': '3',
        'in_stock': '4',
        'recommended': '2',
        'discounted': '8',
        'point_rate_up': '9',
        'not_display_sale_end': '7'
    }
    def __init__(self, row_dict):
        search_query_dict = {}
        for key in self.categories['items']:
            if not row_dict['category']:
                break
            if row_dict['category'] in key:
                search_query_dict[self.categories['keyname']] = self.categories['items'][key]['value']
                if row_dict['related_category']:
                    for rkey in self.categories['items'][key]['related']:
                        if row_dict['related_category'] in rkey:
                            search_query_dict[self.categories['related_keyname']] = self.categories['items'][key]['related'][rkey]
                            break
                break
        search_query_dict['sci'] = '0' # fixed
        search_query_dict['ss'] = '8' # fixed
        search_query_dict['sl'] = '80' # fixed
        search_query_dict['nf'] = '1' # fixed
        search_query_dict['smt'] = row_dict['keyword']
        search_query_dict['ssy'] = row_dict['ssy'] or ''
        search_query_dict['ssm'] = row_dict['ssm'] or ''
        search_query_dict['sey'] = row_dict['sey'] or ''
        search_query_dict['sem'] = row_dict['sem'] or ''
        
        sonota_list = []
        for k in self.sonota:
            if k in row_dict and row_dict[k] == 'y':
                sonota_list.append(self.sonota[k])
        search_query_dict['nd[]'] = sonota_list
        self.search_query_dict = search_query_dict
        self.price_min = row_dict['price_min']
        self.price_max = row_dict['price_max']
        self.toriyose = row_dict['toriyose'] == 'y'
        self.in_stock_now = row_dict['in_stock_now'] == 'y'
        self.can_be_reserved = row_dict['can_be_reserved'] == 'y'

    @property
    def querystring(self):
        return urlencode(self.search_query_dict, True)

    def check_sale_status(self, status_str):
        if status_str == '取り寄せ' and self.toriyose:
            return True
        if status_str == '予約受付中' and self.can_be_reserved:
            return True
        if status_str in ['在庫あり', '残りわずか'] and self.in_stock_now:
            return True
        return False

    def check_price(self, price):
        if self.price_min:
            if price <= int(self.price_min):
                return False
        if self.price_max:
            if price >= int(self.price_max):
                return False
        return True

    def check(self, price, status_str):
        if not self.check_price(price):
            return False
        if not self.check_sale_status(status_str):
            
            return False
        return True
        
class QueryManager(object):
    columns = {
        'do_search': 1,
        'keyword': 2,
        'category': 3,
        'related_category': 4,
        'ssy': 5,
        'ssm': 6,
        'sey': 7,
        'sem': 8,
        'new_arrival': 9,
        'bonus': 10,
        'reservation': 11,
        'in_stock': 12,
        'recommended': 13,
        'discounted': 14,
        'point_rate_up': 15,
        'not_display_sale_end': 16,
        'price_min': 17,
        'price_max': 18,
        'toriyose': 19,
        'in_stock_now':20,
        'can_be_reserved': 21,
    }
   
    query_start_row = 4

    def __init__(self, sheet):
        self.sheet = sheet

    def _get_cell_value(self, r, c_name):
        return self.sheet.cell(r, self.columns[c_name]).value

    @property
    def queries(self):
        queries = []
        r = self.query_start_row
        while True:
            do_search = self._get_cell_value(r, 'do_search')
            if not do_search:
                break
            if do_search != 'y':
                r += 1
                continue
            query = {}
            for k in self.columns:
                query[k] = self._get_cell_value(r, k)
            query_obj = SearchQuery(query)
            queries.append(query_obj)
            r += 1

        return queries

class AnimatePipeline(object):
    excel_file_path = 'アニメイト商品取得ツール.xlsx'
    def open_spider(self, spider):
        try:
            book = openpyxl.load_workbook(self.excel_file_path, data_only = True)
        except Exception as e:
            raise Exception('エクセルファイルが開ませんでした')
        else:
            query_sheet = book.worksheets[0]
            result_sheet = book.worksheets[1]

            query_manager = QueryManager(query_sheet)
            result_manager = ResultManager(result_sheet)
            self.book = book
            spider.query_manager = query_manager
            spider.result_manager = result_manager
            
    def close_spider(self, spider):
        self.book.save(self.excel_file_path)

    def process_item(self, item, spider):
        if spider.result_manager.check_if_already_exists(item):
            spider.logger.info(f'商品:{item["sku"]} はすでにシートに存在します') 
            return
        spider.result_manager.record_item(item)
        spider.logger.info(f'商品:{item["sku"]}を処理しました。')
        return item

class AnimateImagePipeline(ImagesPipeline):
    def get_media_requests(self, item, info):
        if len(item["image_urls"]) < 1:
            return
        yield scrapy.Request(item['image_urls'][0], meta={ 'id': f'{item["sku"]}_thumbnail' })

    def resize(self, item):
        image_data = Image.open(item['thumbnail_path'])
        image_resized = image_data.resize((50, 50), Image.NEAREST)
        image_resized.save(item['thumbnail_path'])

    def item_completed(self, results, item, info):
        file_paths = [x['path'] for ok, x in results if ok]
        if len(file_paths) != 1:
            info.spider.logger.error(f'{item["sku"]}のサムネイル画像が見つかりませんでした。')
        else:
            image_path = f'{info.spider.settings.get("IMAGES_STORE")}{file_paths[0]}'
            item["thumbnail_path"] = image_path
            self.resize(item)
            info.spider.result_manager.embed_thumbnail(item)
            return item 
    def file_path(self, request, response = None, info = None):
        custom_file_name = request.meta.get('id')
        return f'{custom_file_name}.png'
        